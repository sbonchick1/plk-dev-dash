import json
import io
import zipfile
import reimport json
import io
import zipfile
import re
from http.server import BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLOR_MAP = {
    "Prospect":   "E8521A",
    "SA":         "E8521A",
    "PC":         "E8521A",
    "Permitting": "E8521A",
    "UC":         "E8521A",
    "Open":       "E8521A",
    "FY BU":      "7B2D8B",
    "Upside":     "C1272D",
    "Budget":     "00A99D",
}

HEADER_FILL = PatternFill("solid", fgColor="374151")
STRIPE_FILL = PatternFill("solid", fgColor="F3F4F6")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
STATUSES    = ["Prospect", "SA", "PC", "Permitting", "UC", "Open"]


def build_xlsx(payload):
    division_name = payload.get("divisionName", "Division")
    labels  = payload.get("labels", [])
    values  = payload.get("displayValues", [])
    budget  = payload.get("budget", 0)
    fy_bu   = payload.get("fyBU", 0)
    upside  = payload.get("upsideCount", 0)
    gap     = payload.get("gap", 0)
    sites   = payload.get("sites", [])

    bar_labels = STATUSES + ["FY BU", "Upside", "Budget"]
    n = len(bar_labels)

    # Waterfall base (where bar starts) and # SIPs (bar height)
    waterfall_vals = []
    sip_vals       = []
    cumulative = 0
    for lbl in STATUSES:
        idx = labels.index(lbl) if lbl in labels else -1
        v = values[idx] if idx >= 0 else 0
        waterfall_vals.append(cumulative)
        sip_vals.append(v)
        cumulative += v
    waterfall_vals.append(0);      sip_vals.append(fy_bu)    # FY BU from 0
    waterfall_vals.append(fy_bu);  sip_vals.append(upside)   # Upside on top of FY BU
    waterfall_vals.append(0);      sip_vals.append(budget)   # Budget from 0

    wb = Workbook()

    # ══════════════════════════════════════════════════════
    # SHEET 1: Waterfall Chart
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Waterfall Chart"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = f"{division_name} — 2026 Pipeline Waterfall"
    t.font      = Font(bold=True, size=14, color="E8521A")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    # Header row (row 3): Stage | Waterfall | # SIPs
    for col, hdr in enumerate(["Stage", "Waterfall", "# SIPs"], 1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Data rows start at row 4
    table_row = 4
    for i, lbl in enumerate(bar_labels):
        fill = STRIPE_FILL if i % 2 == 0 else WHITE_FILL
        hc   = COLOR_MAP.get(lbl, "6B7280")
        r    = table_row + i

        # Col A: Stage
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font      = Font(bold=(lbl in ("FY BU", "Budget", "Upside")), color=hc)
        c1.fill      = fill
        c1.alignment = Alignment(horizontal="left", vertical="center")

        # Col B: Waterfall (base — where bar starts)
        c2 = ws.cell(row=r, column=2, value=waterfall_vals[i])
        c2.font          = Font(color="9CA3AF")
        c2.fill          = fill
        c2.alignment     = Alignment(horizontal="center", vertical="center")
        c2.number_format = "0"

        # Col C: # SIPs (actual count / bar height)
        c3 = ws.cell(row=r, column=3, value=sip_vals[i])
        c3.font          = Font(bold=True, color=hc)
        c3.fill          = fill
        c3.alignment     = Alignment(horizontal="center", vertical="center")
        c3.number_format = "0"

        ws.row_dimensions[r].height = 17

    table_end = table_row + n - 1  # last data row

    # Summary stats below table
    sr = table_end + 2
    gc = "059669" if gap >= 0 else "DC2626"

    def stat_row(row, label, value, fmt="0", color="374151"):
        ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(bold=True, color=color)
        c.number_format = fmt

    stat_row(sr,   "FY BU",                 fy_bu)
    stat_row(sr+1, "Budget",                budget)
    stat_row(sr+2, "Gap (FY BU vs Budget)", gap,                       "+0;-0;0", gc)
    stat_row(sr+3, "Gap %",                 gap/budget if budget else 0, "0%",    gc)
    stat_row(sr+4, "FY BU + Upside",        fy_bu + upside)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    # ══════════════════════════════════════════════════════
    # Build stacked waterfall chart directly from cols A-C
    # Col A (1) = Stage labels  → X-axis categories
    # Col B (2) = Waterfall     → invisible base series
    # Col C (3) = # SIPs        → visible coloured series
    # Chart placed at E2 — column E is empty and visible
    # ══════════════════════════════════════════════════════
    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "stacked"
    chart.overlap   = 100
    chart.title     = f"{division_name} — 2026 Pipeline"
    chart.width     = 26
    chart.height    = 16
    chart.legend    = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.delete         = True
    chart.x_axis.delete         = False
    chart.x_axis.tickLblPos     = "low"
    chart.x_axis.numFmt         = "General"
    chart.x_axis.axPos          = "b"

    # Series 1 — invisible base (col B, Waterfall)
    base_ref = Reference(ws, min_col=2, min_row=table_row, max_row=table_end)
    base_ser = Series(base_ref, title="base")
    base_ser.graphicalProperties.solidFill      = "FFFFFF"
    base_ser.graphicalProperties.line.solidFill = "FFFFFF"
    base_ser.graphicalProperties.line.width     = 0
    chart.append(base_ser)

    # Series 2 — visible coloured bars (col C, # SIPs)
    bar_ref = Reference(ws, min_col=3, min_row=table_row, max_row=table_end)
    bar_ser = Series(bar_ref, title="# SIPs")
    bar_ser.graphicalProperties.solidFill      = "E8521A"
    bar_ser.graphicalProperties.line.solidFill = "FFFFFF"
    bar_ser.graphicalProperties.line.width     = 6350

    for i, lbl in enumerate(bar_labels):
        dp = DataPoint(idx=i)
        hc = COLOR_MAP.get(lbl, "9CA3AF")
        dp.graphicalProperties.solidFill      = hc
        dp.graphicalProperties.line.solidFill = hc
        bar_ser.dPt.append(dp)

    dl = DataLabelList()
    dl.showVal = True; dl.showCatName = False; dl.showSerName = False
    dl.showPercent = False; dl.showLegendKey = False; dl.position = "outEnd"
    bar_ser.dLbls = dl
    chart.append(bar_ser)

    # X-axis categories from col A
    cats = Reference(ws, min_col=1, min_row=table_row, max_row=table_end)
    chart.set_categories(cats)

    # Place chart at E2 — fully visible, no hidden columns involved
    ws.add_chart(chart, "E2")

    # ══════════════════════════════════════════════════════
    # SHEET 2: Site Detail
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Site Detail")
    ws2.sheet_view.showGridLines = False

    hdrs   = ["SIP ID","Rest No","FZ","Address","City","ST","Status",
              "FZ Proj Open Date","PLK Proj Open Date","Risk Level","Last Comments"]
    widths = [12, 10, 20, 24, 16, 6, 12, 18, 18, 12, 44]

    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 18

    risk_fills = {"low":"D1FAE5","medium":"FEF3C7","high":"FEE2E2","upside":"EDE9FE","2027+":"E0F2FE"}

    for ri, s in enumerate(sites, 2):
        row_vals = [
            s.get("sipId",""),    s.get("restNum",""),   s.get("fz",""),
            s.get("address",""),  s.get("city",""),       s.get("state",""),
            s.get("status",""),   s.get("fzOpenDate",""), s.get("plkOpenDate",""),
            s.get("riskLevel",""),s.get("lastComment",""),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 11))
        ws2.row_dimensions[ri].height = 16
        rf = risk_fills.get(s.get("riskLevel","").strip().lower())
        if rf:
            ws2.cell(row=ri, column=10).fill = PatternFill("solid", fgColor=rf)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"

    # ══════════════════════════════════════════════════════
    # Save → patch chart XML: numRef → strRef for X-axis
    # ══════════════════════════════════════════════════════
    pre_buf = io.BytesIO()
    wb.save(pre_buf)
    pre_buf.seek(0)

    pt_tags = "".join(
        f'<pt idx="{i}"><v>{lbl}</v></pt>'
        for i, lbl in enumerate(bar_labels)
    )
    str_cache = f'<strCache><ptCount val="{n}"/>{pt_tags}</strCache>'

    out_buf = io.BytesIO()
    with zipfile.ZipFile(pre_buf, "r") as zin, \
         zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/charts/chart1.xml":
                xml = data.decode("utf-8")

                def _patch_cat(m):
                    inner = m.group(1)
                    f_match = re.search(r'<f>(.*?)</f>', inner, re.DOTALL)
                    f_tag = f_match.group(0) if f_match else ""
                    return f'<cat><strRef>{f_tag}{str_cache}</strRef></cat>'

                xml = re.sub(
                    r'<cat>\s*<numRef>(.*?)</numRef>\s*</cat>',
                    _patch_cat, xml, flags=re.DOTALL
                )
                data = xml.encode("utf-8")
            zout.writestr(item, data)

    out_buf.seek(0)
    return out_buf.read()


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_POST(self):
        try:
            length  = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length))
        except Exception as e:
            self._respond_error(400, f"Bad request: {e}"); return
        try:
            xlsx_bytes = build_xlsx(payload)
        except Exception as e:
            self._respond_error(500, str(e)); return

        division_name = payload.get("divisionName", "Division")
        safe     = "".join(c if c.isalnum() or c in " _-" else "_" for c in division_name)
        filename = f"Waterfall_{safe}.xlsx"

        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(xlsx_bytes)))
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def _respond_error(self, code, msg):
        body = json.dumps({"error": msg}).encode()
        self.send_response(code)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args):
        pass
from http.server import BaseHTTPRequestHandler
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLOR_MAP = {
    "Prospect":   "E8521A",
    "SA":         "E8521A",
    "PC":         "E8521A",
    "Permitting": "E8521A",
    "UC":         "E8521A",
    "Open":       "E8521A",
    "FY BU":      "7B2D8B",
    "Upside":     "C1272D",
    "Budget":     "00A99D",
}

HEADER_FILL = PatternFill("solid", fgColor="374151")
STRIPE_FILL = PatternFill("solid", fgColor="F3F4F6")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
STATUSES    = ["Prospect", "SA", "PC", "Permitting", "UC", "Open"]


def build_xlsx(payload):
    division_name = payload.get("divisionName", "Division")
    labels  = payload.get("labels", [])          # ["Prospect","SA",…,"FY BU","Upside","Budget"]
    values  = payload.get("displayValues", [])   # raw bar heights (individual counts)
    budget  = payload.get("budget", 0)
    fy_bu   = payload.get("fyBU", 0)
    upside  = payload.get("upsideCount", 0)
    gap     = payload.get("gap", 0)
    sites   = payload.get("sites", [])

    bar_labels = STATUSES + ["FY BU", "Upside", "Budget"]
    n = len(bar_labels)

    # ── Build Waterfall (base) and # SIPs (count) for each stage ─────────────
    # Waterfall = where the bar starts (the invisible lift value)
    # # SIPs    = the actual count for that stage (bar height)
    #
    # Rules:
    #   Prospect → starts at 0
    #   SA       → starts at sum(Prospect)
    #   PC       → starts at sum(Prospect + SA)
    #   ...each status builds on previous cumulative
    #   FY BU    → starts at 0 (standalone)
    #   Upside   → starts at FY BU (stacks on top of FY BU)
    #   Budget   → starts at 0 (standalone)

    waterfall_vals = []   # invisible base (where bar starts)
    sip_vals       = []   # visible bar height (# SIPs / count)

    cumulative = 0
    for lbl in STATUSES:
        idx = labels.index(lbl) if lbl in labels else -1
        v = values[idx] if idx >= 0 else 0
        waterfall_vals.append(cumulative)   # starts where previous ended
        sip_vals.append(v)
        cumulative += v

    # FY BU — starts at 0
    waterfall_vals.append(0)
    sip_vals.append(fy_bu)

    # Upside — starts at FY BU (stacks on top)
    waterfall_vals.append(fy_bu)
    sip_vals.append(upside)

    # Budget — starts at 0
    waterfall_vals.append(0)
    sip_vals.append(budget)

    wb = Workbook()

    # ══════════════════════════════════════════════════════
    # SHEET 1: Waterfall Chart
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Waterfall Chart"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = f"{division_name} — 2026 Pipeline Waterfall"
    t.font      = Font(bold=True, size=14, color="E8521A")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 6

    # ── Table header: Stage | Waterfall | # SIPs ──────────
    headers = ["Stage", "Waterfall", "# SIPs"]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=hdr)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # ── Data rows ─────────────────────────────────────────
    table_row = 4
    for i, lbl in enumerate(bar_labels):
        fill = STRIPE_FILL if i % 2 == 0 else WHITE_FILL
        hc   = COLOR_MAP.get(lbl, "6B7280")
        wf   = waterfall_vals[i]   # where bar starts
        sip  = sip_vals[i]         # bar height / count

        r = table_row + i

        # Col A: Stage
        c1 = ws.cell(row=r, column=1, value=lbl)
        c1.font      = Font(bold=(lbl in ("FY BU", "Budget", "Upside")), color=hc)
        c1.fill      = fill
        c1.alignment = Alignment(horizontal="left", vertical="center")

        # Col B: Waterfall (base / where bar starts)
        c2 = ws.cell(row=r, column=2, value=wf)
        c2.font          = Font(color="9CA3AF")   # grey — this is metadata not a headline number
        c2.fill          = fill
        c2.alignment     = Alignment(horizontal="center", vertical="center")
        c2.number_format = "0"

        # Col C: # SIPs (actual count)
        c3 = ws.cell(row=r, column=3, value=sip)
        c3.font          = Font(bold=True, color=hc)
        c3.fill          = fill
        c3.alignment     = Alignment(horizontal="center", vertical="center")
        c3.number_format = "0"

        ws.row_dimensions[r].height = 17

    table_end = table_row + n - 1

    # ── Summary stats ─────────────────────────────────────
    sr = table_end + 2
    gc = "059669" if gap >= 0 else "DC2626"

    def stat_row(row, label, value, fmt="0", color="374151"):
        ws.cell(row=row, column=1, value=label).font = Font(color="6B7280")
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(bold=True, color=color)
        c.number_format = fmt
        # Also show in col C for clarity
        c3 = ws.cell(row=row, column=3, value=value)
        c3.font = Font(color="9CA3AF")
        c3.number_format = fmt

    stat_row(sr,   "FY BU",                 fy_bu)
    stat_row(sr+1, "Budget",                budget)
    stat_row(sr+2, "Gap (FY BU vs Budget)", gap,                  "+0;-0;0", gc)
    stat_row(sr+3, "Gap %",                 gap/budget if budget else 0, "0%", gc)
    stat_row(sr+4, "FY BU + Upside",        fy_bu + upside)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10

    # ══════════════════════════════════════════════════════
    # Chart data block — hidden columns E/F/G
    # E = labels, F = waterfall base, G = SIP count
    # ══════════════════════════════════════════════════════
    LABEL_COL = 5   # E
    BASE_COL  = 6   # F  (invisible lift)
    BAR_COL   = 7   # G  (visible bar = # SIPs)

    chart_row_start = 3
    chart_row_end   = chart_row_start + n - 1

    for i, lbl in enumerate(bar_labels):
        r = chart_row_start + i
        ws.cell(row=r, column=LABEL_COL, value=lbl)
        ws.cell(row=r, column=BASE_COL,  value=waterfall_vals[i])
        ws.cell(row=r, column=BAR_COL,   value=sip_vals[i])

    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True

    # ══════════════════════════════════════════════════════
    # Build stacked waterfall chart
    # Series 1: invisible base (lifts bars to correct position)
    # Series 2: visible coloured bars (# SIPs)
    # ══════════════════════════════════════════════════════
    chart = BarChart()
    chart.type      = "col"
    chart.grouping  = "stacked"
    chart.overlap   = 100
    chart.title     = f"{division_name} — 2026 Pipeline"
    chart.width     = 26
    chart.height    = 16
    chart.legend    = None
    chart.y_axis.majorGridlines = None
    chart.y_axis.delete         = True
    chart.x_axis.delete         = False
    chart.x_axis.tickLblPos     = "low"
    chart.x_axis.numFmt         = "General"
    chart.x_axis.axPos          = "b"

    # Series 1 — invisible base (white)
    base_ref = Reference(ws, min_col=BASE_COL, min_row=chart_row_start, max_row=chart_row_end)
    base_ser = Series(base_ref, title="base")
    base_ser.graphicalProperties.solidFill      = "FFFFFF"
    base_ser.graphicalProperties.line.solidFill = "FFFFFF"
    base_ser.graphicalProperties.line.width     = 0
    chart.append(base_ser)

    # Series 2 — visible coloured bars
    bar_ref = Reference(ws, min_col=BAR_COL, min_row=chart_row_start, max_row=chart_row_end)
    bar_ser = Series(bar_ref, title="# SIPs")
    bar_ser.graphicalProperties.solidFill      = "E8521A"
    bar_ser.graphicalProperties.line.solidFill = "FFFFFF"
    bar_ser.graphicalProperties.line.width     = 6350

    # Colour each bar individually
    for i, lbl in enumerate(bar_labels):
        dp = DataPoint(idx=i)
        hc = COLOR_MAP.get(lbl, "9CA3AF")
        dp.graphicalProperties.solidFill      = hc
        dp.graphicalProperties.line.solidFill = hc
        bar_ser.dPt.append(dp)

    # Value labels above bars (show # SIPs)
    dl = DataLabelList()
    dl.showVal = True; dl.showCatName = False; dl.showSerName = False
    dl.showPercent = False; dl.showLegendKey = False; dl.position = "outEnd"
    bar_ser.dLbls = dl
    chart.append(bar_ser)

    # X-axis categories
    cats = Reference(ws, min_col=LABEL_COL, min_row=chart_row_start, max_row=chart_row_end)
    chart.set_categories(cats)

    # Place chart to the right of the table
    ws.add_chart(chart, "E2")

    # ══════════════════════════════════════════════════════
    # SHEET 2: Site Detail
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Site Detail")
    ws2.sheet_view.showGridLines = False

    hdrs   = ["SIP ID","Rest No","FZ","Address","City","ST","Status",
              "FZ Proj Open Date","PLK Proj Open Date","Risk Level","Last Comments"]
    widths = [12, 10, 20, 24, 16, 6, 12, 18, 18, 12, 44]

    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 18

    risk_fills = {"low":"D1FAE5","medium":"FEF3C7","high":"FEE2E2","upside":"EDE9FE","2027+":"E0F2FE"}

    for ri, s in enumerate(sites, 2):
        row_vals = [
            s.get("sipId",""),    s.get("restNum",""),   s.get("fz",""),
            s.get("address",""),  s.get("city",""),       s.get("state",""),
            s.get("status",""),   s.get("fzOpenDate",""), s.get("plkOpenDate",""),
            s.get("riskLevel",""),s.get("lastComment",""),
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 11))
        ws2.row_dimensions[ri].height = 16
        rf = risk_fills.get(s.get("riskLevel","").strip().lower())
        if rf:
            ws2.cell(row=ri, column=10).fill = PatternFill("solid", fgColor=rf)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"

    # ══════════════════════════════════════════════════════
    # Save → patch chart XML: numRef → strRef for X-axis labels
    # ══════════════════════════════════════════════════════
    pre_buf = io.BytesIO()
    wb.save(pre_buf)
    pre_buf.seek(0)

    pt_tags = "".join(
        f'<pt idx="{i}"><v>{lbl}</v></pt>'
        for i, lbl in enumerate(bar_labels)
    )
    str_cache = f'<strCache><ptCount val="{n}"/>{pt_tags}</strCache>'

    out_buf = io.BytesIO()
    with zipfile.ZipFile(pre_buf, "r") as zin, \
         zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/charts/chart1.xml":
                xml = data.decode("utf-8")

                def _patch_cat(m):
                    inner = m.group(1)
                    f_match = re.search(r'<f>(.*?)</f>', inner, re.DOTALL)
                    f_tag = f_match.group(0) if f_match else ""
                    return f'<cat><strRef>{f_tag}{str_cache}</strRef></cat>'

                xml = re.sub(
                    r'<cat>\s*<numRef>(.*?)</numRef>\s*</cat>',
                    _patch_cat, xml, flags=re.DOTALL
                )
                data = xml.encode("utf-8")
            zout.writestr(item, data)

    out_buf.seek(0)
    return out_buf.read()


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_POST(self):
        try:
            length  = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length))
        except Exception as e:
            self._respond_error(400, f"Bad request: {e}"); return
        try:
            xlsx_bytes = build_xlsx(payload)
        except Exception as e:
            self._respond_error(500, str(e)); return

        division_name = payload.get("divisionName", "Division")
        safe     = "".join(c if c.isalnum() or c in " _-" else "_" for c in division_name)
        filename = f"Waterfall_{safe}.xlsx"

        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(xlsx_bytes)))
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def _respond_error(self, code, msg):
        body = json.dumps({"error": msg}).encode()
        self.send_response(code)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args):
        pass
